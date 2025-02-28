from datetime import datetime, timedelta
from typing import Any, Optional, List
from openpyxl.styles import PatternFill

def parse_datetime(value: Any, fmt: str = "%Y-%m-%d") -> Optional[datetime]:
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, str):
        try:
            return datetime.fromisoformat(value)
        except ValueError:
            try:
                return datetime.strptime(value, fmt)
            except ValueError:
                return None
    return None

def format_timedelta(delta: timedelta) -> str:
    total_seconds = abs(delta.total_seconds())
    hours = int(total_seconds // 3600)
    minutes = int((total_seconds % 3600) // 60)
    sign = '-' if delta.total_seconds() < 0 else '+'
    return f"{sign}{hours:02d}:{minutes:02d}"

def _score_delta(delta: timedelta) -> int:
    minutes = abs(delta.total_seconds()) / 60
    if minutes < 20:
        return 0
    elif minutes < 60:
        return 1
    elif minutes < 120:
        return 2
    else:
        return 3

def parse_time_diff(diff_str: Optional[str]) -> timedelta:
    if not diff_str:
        return timedelta(0)
    sign = -1 if diff_str.startswith('-') else 1
    hours, minutes = map(int, diff_str[1:].split(':'))
    return timedelta(hours=hours, minutes=minutes) * sign

def calculate_orthodox_easter(year: int) -> datetime.date:
    hl_year = year - 2
    remainder = (hl_year % 19) * 11
    epact = remainder % 30
    easter_day = 44 - epact
    if epact > 23:
        easter_date = datetime(year, 4, int(easter_day))
    else:
        first_march = datetime(year, 3, 1)
        easter_date = first_march + timedelta(days=easter_day - 1)
    easter_date = easter_date + timedelta(days=13)  # add 13 days
    easter_date = easter_date + timedelta(days=1)   # add 1 day
    offset = (6 - easter_date.weekday()) % 7
    easter_date = easter_date + timedelta(days=offset)
    return easter_date.date()

def get_holidays(year: int) -> set:
    easter = calculate_orthodox_easter(year)
    return {
        datetime(year, 1, 1).date(),
        datetime(year, 1, 6).date(),
        datetime(year, 3, 25).date(),
        datetime(year, 5, 1).date(),
        datetime(year, 8, 15).date(),
        datetime(year, 10, 28).date(),
        datetime(year, 12, 25).date(),
        datetime(year, 12, 26).date(),
        (easter - timedelta(days=48)),
        (easter - timedelta(days=2)),
        easter,
        (easter + timedelta(days=1))
    }

def sort_workers(workers: List[Any]) -> List[Any]:
    return sorted(workers, key=lambda w: (w.last_name, w.first_name, w.afm, w.date))

# Define fill styles that might be used by various modules
grey_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")      # weekend only
holiday_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")    # holiday only (gold)
both_fill = PatternFill(start_color="FF4500", end_color="FF4500", fill_type="solid")       # both holiday and weekend (orange-red)

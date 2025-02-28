from dataclasses import dataclass, field
from datetime import datetime, timedelta, time
from typing import Optional, List, Any, Dict
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from utils import (
    parse_datetime,
    format_timedelta,
    _score_delta,
    parse_time_diff,
    get_holidays,
    sort_workers,
    grey_fill,
    holiday_fill,
    both_fill
)

# --- New Helper Functions for Premium Hours ---

def compute_overlap(interval_start: datetime, interval_end: datetime, window_start: datetime, window_end: datetime) -> float:
    """Return the overlap (in hours) between an interval and a time window."""
    latest_start = max(interval_start, window_start)
    earliest_end = min(interval_end, window_end)
    delta = (earliest_end - latest_start).total_seconds()
    return max(0, delta) / 3600

def calculate_night_hours(planned_start: datetime, planned_end: datetime) -> float:
    """
    Calculates the number of hours in the interval [planned_start, planned_end]
    that fall within the night period (22:00 to 06:00).
    """
    total_night = 0.0
    day = planned_start.date()
    last_day = planned_end.date()
    while day <= last_day:
        night_start = datetime.combine(day, time(22, 0))
        night_end = datetime.combine(day + timedelta(days=1), time(6, 0))
        total_night += compute_overlap(planned_start, planned_end, night_start, night_end)
        day += timedelta(days=1)
    return total_night

def calculate_sunday_hours(planned_start: datetime, planned_end: datetime) -> float:
    """
    Calculates the number of hours in the interval [planned_start, planned_end]
    that fall on Sunday (00:00 to 23:59:59).
    """
    total_sunday = 0.0
    day = planned_start.date()
    last_day = planned_end.date()
    while day <= last_day:
        if day.weekday() == 6:  # Sunday (Monday=0, Sunday=6)
            day_start = datetime.combine(day, time(0, 0))
            day_end = datetime.combine(day, time(23, 59, 59))
            total_sunday += compute_overlap(planned_start, planned_end, day_start, day_end)
        day += timedelta(days=1)
    return total_sunday

# --- Data Model ---

@dataclass
class Worker:
    branch: str
    afm: str
    last_name: str
    first_name: str
    date: datetime
    break_time: Any
    work: str
    planned_start: Optional[datetime] = None
    planned_end: Optional[datetime] = None
    actual_start: Optional[datetime] = None
    actual_end: Optional[datetime] = None
    start_diff: Optional[str] = None
    finish_diff: Optional[str] = None
    summary: Optional[str] = None
    severity: Optional[str] = None
    worker_score: int = 0

    SPECIAL_VALUES: List[str] = field(default_factory=lambda: ["REST/DAY OFF", "NON-WORKING", "UNPLANNED"])

    @property
    def is_special(self) -> bool:
        return self.work in self.SPECIAL_VALUES

    def validate_times(self):
        if not self.is_special:
            if not (self.planned_start and self.planned_end):
                raise ValueError("Invalid planned shift times")
            if self.planned_start >= self.planned_end:
                raise ValueError("Planned end time must be after start time")

    def calculate_time_differences(self):
        if self.is_special:
            return
        if self.planned_start and self.actual_start:
            delta = self.planned_start - self.actual_start
            self.start_diff = format_timedelta(delta)
        if self.planned_end and self.actual_end:
            # Calculate raw delta between actual and planned end
            delta = self.actual_end - self.planned_end
            # If break_time is "Outside NUM", subtract NUM minutes from the delta
            if isinstance(self.break_time, str) and self.break_time.startswith("Outside"):
                try:
                    break_minutes = int(self.break_time.split()[1])
                    delta = delta - timedelta(minutes=break_minutes)
                except Exception as e:
                    print(f"Error processing break time for AFM {self.afm}: {e}")
            self.finish_diff = format_timedelta(delta)

# --- Data Processing Functions ---

def process_planned_data(planned_entries: List[Dict]) -> List[Worker]:
    workers = []
    for entry in planned_entries:
        try:
            afm = str(entry.get('afm', '')).strip()
            if not afm:
                raise ValueError("Missing AFM")
            work = entry.get('work', '')
            planned_start = parse_datetime(entry.get('clock_in')) if work not in ["REST/DAY OFF", "NON-WORKING"] else None
            planned_end = parse_datetime(entry.get('clock_out')) if work not in ["REST/DAY OFF", "NON-WORKING"] else None
            date_raw = parse_datetime(entry.get('date'))
            if not date_raw:
                raise ValueError("Invalid date")
            date_val = date_raw.date()
            branch_raw = entry.get('branch')
            branch = "Unknown" if branch_raw is None or branch_raw == "" else str(branch_raw)
            # Do not modify planned_end for premium calculations.
            worker = Worker(
                branch=branch,
                afm=afm,
                last_name=entry.get('last_name', ''),
                first_name=entry.get('first_name', ''),
                date=datetime.combine(date_val, datetime.min.time()),
                planned_start=planned_start,
                planned_end=planned_end,
                break_time=entry.get('break', ''),
                work=work
            )
            worker.validate_times()
            workers.append(worker)
        except Exception as e:
            print(f"Skipping planned entry: {e}")
    return workers

def process_actual_times(workers: List[Worker], actual_entries: List[Dict]) -> None:
    for entry in actual_entries:
        try:
            afm = str(entry.get('afm', '')).strip()
            if not afm:
                continue
            actual_start = parse_datetime(entry.get('clock_in'))
            actual_end = parse_datetime(entry.get('clock_out'))
            if not (actual_start or actual_end):
                continue
            target_date = (actual_start or actual_end).date()
            for worker in workers:
                if worker.afm == afm and worker.date.date() == target_date:
                    if actual_start:
                        worker.actual_start = actual_start
                    if actual_end:
                        worker.actual_end = actual_end
                    if not worker.is_special:
                        worker.calculate_time_differences()
                    break
        except Exception as e:
            print(f"Error processing actual entry: {e}")

def add_summary_and_severity(workers: List[Worker], remove_future: bool = False, remove_dayoff: bool = False) -> List[Worker]:
    today = datetime.now().date()
    filtered = []
    for worker in workers:
        # If the worker is "special" (i.e. non‐working) then mark as scheduled day off.
        if worker.is_special:
            worker.summary = "Scheduled day off"
            worker.severity = "N/A"
            if remove_dayoff:
                continue  # Skip this worker if we want to remove day off rows
        # If the worker's shift is in the future, mark it.
        elif worker.date.date() > today:
            worker.summary = "Future shift"
            worker.severity = "N/A"
            if remove_future:
                continue  # Skip this worker if we want to remove future shifts
        else:
            score = 0
            parts = []
            start_delta = parse_time_diff(worker.start_diff or "")
            finish_delta = parse_time_diff(worker.finish_diff or "")
            if not (worker.actual_start or worker.actual_end):
                parts.append("Didn't work")
                score = 3
            else:
                if not worker.actual_start:
                    parts.append("Missing clock-in")
                    score += 2 + _score_delta(start_delta)
                if not worker.actual_end:
                    parts.append("Missing clock-out")
                    score += 2 + _score_delta(finish_delta)
                if worker.actual_start and worker.actual_end:
                    total_diff = start_delta + finish_delta
                    parts.append(f"Total diff: {format_timedelta(total_diff)}")
                    score += _score_delta(total_diff)
            worker.worker_score = score
            if score >= 3:
                worker.severity = "Critical"
            elif score == 2:
                worker.severity = "High"
            elif score == 1:
                worker.severity = "Medium"
            else:
                worker.severity = "Low"
            worker.summary = " | ".join(parts)
        filtered.append(worker)
    return filtered


# --- New Function: Calculate Planned Premium Hours ---
def calculate_planned_premium_hours(worker: Worker) -> Dict[str, float]:
    """
    Calculate, from the planned clock in/out (without break adjustments):
      - Total planned hours,
      - Night hours (22:00–06:00; eligible for a 25% premium),
      - Sunday hours (all hours on Sunday; eligible for a 75% premium).
    Note: Hours may overlap (e.g., a Sunday night counts for both).
    """
    if not (worker.planned_start and worker.planned_end):
        return {"total_hours": 0.0, "night_hours": 0.0, "sunday_hours": 0.0}
    total = (worker.planned_end - worker.planned_start).total_seconds() / 3600
    night = calculate_night_hours(worker.planned_start, worker.planned_end)
    sunday = calculate_sunday_hours(worker.planned_start, worker.planned_end)
    return {"total_hours": total, "night_hours": night, "sunday_hours": sunday}

# --- Export to Excel (with sorting, additional columns, and row fills) ---
def worker_to_dict(worker: Worker) -> Dict:
    premium = calculate_planned_premium_hours(worker)
    base = {
        "AFM": worker.afm,
        "Branch": worker.branch,
        "Last Name": worker.last_name,
        "First Name": worker.first_name,
        "Planned Start (ISO)": worker.planned_start.isoformat() if worker.planned_start else "",
        "Actual Start (ISO)": worker.actual_start.isoformat() if worker.actual_start else "",
        "Planned End (ISO)": worker.planned_end.isoformat() if worker.planned_end else "",
        "Actual End (ISO)": worker.actual_end.isoformat() if worker.actual_end else "",
        "Break": worker.break_time,
        "Schedule Start Date": worker.planned_start.strftime("%Y-%m-%d") if worker.planned_start else "",
        "Schedule Start Time": worker.planned_start.strftime("%H:%M") if worker.planned_start else "",
        "Schedule End Date": worker.planned_end.strftime("%Y-%m-%d") if worker.planned_end else "",
        "Schedule End Time": worker.planned_end.strftime("%H:%M") if worker.planned_end else "",
        "Actual Start Date": worker.actual_start.strftime("%Y-%m-%d") if worker.actual_start else "",
        "Actual Start Time": worker.actual_start.strftime("%H:%M") if worker.actual_start else "",
        "Actual End Date": worker.actual_end.strftime("%Y-%m-%d") if worker.actual_end else "",
        "Actual End Time": worker.actual_end.strftime("%H:%M") if worker.actual_end else "",
        "Start Diff": worker.start_diff or "",
        "Finish Diff": worker.finish_diff or "",
        "Summary": worker.summary or "",
        "Severity": worker.severity or ""
    }
    # Add new premium columns:
    base["Planned Hours"] = round(premium["total_hours"], 2)
    base["Planned Night Hours"] = round(premium["night_hours"], 2)
    base["Planned Sunday Hours"] = round(premium["sunday_hours"], 2)
    return base

def export_results(workers: List[Worker], output_path: str) -> None:
    sorted_workers = sort_workers(workers)
    # Define allowed headers for export (exclude the ISO times)
    allowed_headers = [
        "AFM", "Branch", "Last Name", "First Name",
        "Break", "Schedule Start Date", "Schedule Start Time",
        "Schedule End Date", "Schedule End Time",
        "Actual Start Date", "Actual Start Time",
        "Actual End Date", "Actual End Time",
        "Start Diff", "Finish Diff", "Summary", "Severity",
        "Planned Hours", "Planned Night Hours", "Planned Sunday Hours"
    ]
    wb = Workbook()
    ws = wb.active
    ws.append(allowed_headers)
    for w in sorted_workers:
        # Get full dict; then only keep allowed keys for export
        d = worker_to_dict(w)
        row = [d.get(key, "") for key in allowed_headers]
        ws.append(row)
    # Example: color rows based on date (unchanged code)
    date_columns = [6, 8, 10, 12]  # update these if needed
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        has_holiday = False
        has_weekend = False
        for col in date_columns:
            cell = row[col - 1]
            if cell.value:
                try:
                    dt = datetime.strptime(cell.value, "%Y-%m-%d").date()
                    hols = get_holidays(dt.year)
                    if dt in hols:
                        has_holiday = True
                    if dt.weekday() >= 5:
                        has_weekend = True
                except Exception:
                    pass
        if has_holiday and has_weekend:
            row_fill = both_fill
        elif has_holiday:
            row_fill = holiday_fill
        elif has_weekend:
            row_fill = grey_fill
        else:
            row_fill = None
        if row_fill:
            for cell in row:
                cell.fill = row_fill
    wb.save(output_path)

def combine(planned_entries: List[Dict], actual_entries: List[Dict], remove_future: bool, remove_dayoff: bool, output_path: str = None) -> List[Dict]:
    workers = process_planned_data(planned_entries)
    process_actual_times(workers, actual_entries)
    workers = add_summary_and_severity(workers, remove_future, remove_dayoff)
    if output_path:
        export_results(workers, output_path)
    return [worker_to_dict(w) for w in workers]

from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime, timedelta
from collections import defaultdict
from dataclasses import dataclass

@dataclass
class EmployeeRecord:
    afm: str
    last_name: str
    first_name: str
    clock_in: str = None
    clock_out: str = None
    clock_in_dt: datetime = None
    clock_out_dt: datetime = None

class DateTimeParser:
    @staticmethod
    def parse_date(raw) -> datetime.date:
        if isinstance(raw, datetime):
            return datetime(raw.year, raw.day, raw.month).date()
        if isinstance(raw, str):
            try:
                return datetime.strptime(raw.strip(), '%d/%m/%Y').date()
            except ValueError:
                return None
        return None

    @staticmethod
    def parse_time(raw) -> datetime.time:
        if isinstance(raw, datetime):
            return raw.time()
        if isinstance(raw, str) and ':' in raw:
            try:
                return datetime.strptime(raw.strip(), '%H:%M').time()
            except ValueError:
                return None
        return None

    @classmethod
    def to_iso_datetime(cls, date_obj, time_obj) -> str:
        if not date_obj or not time_obj:
            return None
        return datetime.combine(date_obj, time_obj).isoformat()

class ExcelExporter:
    @staticmethod
    def save(data: list[dict], output_path: str) -> None:
        """Saves processed data to Excel file in standard format"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Processed Timetable"
        
        # Write header
        headers = ["afm", "last_name", "first_name", "clock_in", "clock_out"]
        ws.append(headers)
        
        # Write data rows
        for record in data:
            ws.append([
                record['afm'],
                record['last_name'],
                record['first_name'],
                record['clock_in'],
                record['clock_out']
            ])
        
        wb.save(output_path)

class TimetableProcessor:
    def __init__(self, file_path):
        self.file_path = file_path
        self.parser = DateTimeParser()
    
    def process(self) -> list[EmployeeRecord]:
        raw_data = self._load_raw_data()
        records = self._create_records(raw_data)
        complete, problem = self._separate_records(records)
        matched, unmatched = self._match_problem_records(problem)
        return self._prepare_final_output(complete + matched + unmatched)

    def process_and_save(self, output_path: str = None) -> list[dict]:
        """Process data and optionally save to Excel file"""
        processed_data = self.process()
        
        if output_path:
            ExcelExporter.save(processed_data, output_path)
        
        return processed_data

    def _load_raw_data(self) -> list[tuple]:
        workbook = load_workbook(self.file_path)
        sheet = workbook.active
        return [row for row in sheet.iter_rows(min_row=2, values_only=True) if any(row)]

    def _create_records(self, raw_data) -> list[EmployeeRecord]:
        records = []
        for row in raw_data:
            row = list(row) + [None] * (7 - len(row))
            afm = str(row[1]).strip("'") if row[1] else None
            
            date_obj = self.parser.parse_date(row[4])
            time_in = self.parser.parse_time(row[5])
            time_out = self.parser.parse_time(row[6])

            records.append(EmployeeRecord(
                afm=afm,
                last_name=row[2],
                first_name=row[3],
                clock_in=self.parser.to_iso_datetime(date_obj, time_in),
                clock_out=self.parser.to_iso_datetime(date_obj, time_out),
                clock_in_dt=datetime.fromisoformat(self.parser.to_iso_datetime(date_obj, time_in)) 
                    if date_obj and time_in else None,
                clock_out_dt=datetime.fromisoformat(self.parser.to_iso_datetime(date_obj, time_out)) 
                    if date_obj and time_out else None
            ))
        return records

    def _separate_records(self, records) -> tuple[list, list]:
        complete, problem = [], []
        for r in records:
            (complete if r.clock_in and r.clock_out else problem).append(r)
        return complete, problem

    def _match_problem_records(self, problem_records) -> tuple[list, list]:
        matcher = RecordMatcher()
        return matcher.match(problem_records)

    def _prepare_final_output(self, records) -> list[dict]:
        seen = set()
        unique = []
        for r in sorted(records, key=lambda x: (x.afm, x.clock_in or "")):
            key = (r.afm, r.last_name, r.first_name, r.clock_in, r.clock_out)
            if key not in seen:
                seen.add(key)
                unique.append({
                    'afm': r.afm,
                    'last_name': r.last_name,
                    'first_name': r.first_name,
                    'clock_in': r.clock_in,
                    'clock_out': r.clock_out
                })
        return unique

class RecordMatcher:
    def match(self, records) -> tuple[list, list]:
        self._mark_unused(records)
        employee_groups = self._group_by_employee(records)
        return self._process_matches(employee_groups, records)

    def _mark_unused(self, records):
        for r in records:
            r.used = False

    def _group_by_employee(self, records) -> defaultdict:
        groups = defaultdict(lambda: {'ins': [], 'outs': []})
        for r in records:
            key = (r.afm, r.last_name, r.first_name)
            if r.clock_in and not r.clock_out:
                groups[key]['ins'].append(r)
            elif r.clock_out and not r.clock_in:
                groups[key]['outs'].append(r)
        return groups

    def _process_matches(self, groups, original_records) -> tuple[list, list]:
        matched = []
        for emp_key, entries in groups.items():
            sorted_ins = sorted(
                [r for r in entries['ins'] if r.clock_in_dt],
                key=lambda x: x.clock_in_dt
            )
            sorted_outs = sorted(
                [r for r in entries['outs'] if r.clock_out_dt],
                key=lambda x: x.clock_out_dt
            )
            matched += self._find_valid_matches(sorted_ins, sorted_outs)

        unmatched = [r for r in original_records if not getattr(r, 'used', False)]
        return matched, unmatched

    def _find_valid_matches(self, ins, outs) -> list:
        matched = []
        i = j = 0
        while i < len(ins) and j < len(outs):
            in_rec, out_rec = ins[i], outs[j]
            if self._valid_match(in_rec.clock_in_dt, out_rec.clock_out_dt):
                in_rec.used = out_rec.used = True
                matched.append(self._create_combined_record(in_rec, out_rec))
                i += 1
                j += 1
            elif in_rec.clock_in_dt.date() < out_rec.clock_out_dt.date():
                i += 1
            else:
                j += 1
        return matched

    def _valid_match(self, clock_in, clock_out) -> bool:
        """
        Returns True only if:
        1) clock_in and clock_out are valid datetimes,
        2) The total shift time is non-negative and <= 12 hours,
        3) They occur on the same day or adjacent days.
        """
        if not clock_in or not clock_out:
            return False
        
        shift_seconds = (clock_out - clock_in).total_seconds()
        # If negative or over 12 hours => invalid
        if shift_seconds < 0 or shift_seconds > 43_200:  # 12 hours = 43,200 seconds
            return False

        date_diff = (clock_out.date() - clock_in.date()).days
        # Original logic allowed same day or next day
        return (date_diff == 0) or (date_diff == 1)

    def _create_combined_record(self, in_rec, out_rec) -> EmployeeRecord:
        return EmployeeRecord(
            afm=in_rec.afm,
            last_name=in_rec.last_name,
            first_name=in_rec.first_name,
            clock_in=in_rec.clock_in,
            clock_out=out_rec.clock_out,
            clock_in_dt=in_rec.clock_in_dt,
            clock_out_dt=out_rec.clock_out_dt
        )

def process_timetable(file_path: str) -> list[dict]:
    processor = TimetableProcessor(file_path)
    output_path = file_path
    return processor.process_and_save(output_path)